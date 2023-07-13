from openpyxl import load_workbook

excel_path = r"C:\Users\USER\PycharmProjects\pythonProject3\전산 프로젝트\수료증 자동생성 후 PDF 변환\수료명단.xlsx"

wb = load_workbook(excel_path, data_only=True)
ws = wb.active

name_list = []
birthday_list = []
ho_list = []
for i in range(2, ws.max_row + 1):
    name_list.append(ws.cell(i, 1).value)
    birthday_list.append(ws.cell(i, 2).value)
    ho_list.append(ws.cell(i, 3).value)

print("이름: ", name_list)
print("생일:", birthday_list)
print("호:", ho_list)
