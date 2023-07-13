import glob
from openpyxl import load_workbook
from openpyxl import Workbook

판매보고들 = glob.glob(r"C:\Users\USER\PycharmProjects\pythonProject3\전산 프로젝트\여러개의 엑셀 판매 보고서 합치기\판매보고_*.xlsx")

판매점 = []
날짜 = []
금액 = []

for 판매보고 in 판매보고들:
    wb = load_workbook(판매보고, data_only=True)
    ws = wb.active
    판매점.append(ws['B1'].value)
    날짜.append(ws['B2'].value)
    금액.append(ws['B3'].value)

print(판매점)
print(날짜)
print(금액)

try:
    wb = load_workbook(r"C:\Users\USER\PycharmProjects\pythonProject3\전산 프로젝트\여러개의 엑셀 판매 보고서 합치기\결과.xlsx", data_only=True)
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active

for i in range(len(판매점)):
    ws.cell(row=i + 1, column=1).value = 판매점[i]
    ws.cell(row=i + 1, column=2).value = 날짜[i]
    ws.cell(row=i + 1, column=3).value = 금액[i]

wb.save(r"C:\Users\USER\PycharmProjects\pythonProject3\전산 프로젝트\여러개의 엑셀 판매 보고서 합치기\결과.xlsx")
